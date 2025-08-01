#!/usr/bin/env python3
"""
Comprehensive Employment Variable Analysis Script

This script combines the functionality of all three employment analysis scripts:
1. Accurate analysis using official Bay Area Metro TM2 documentation
2. Discovery of undocumented employment variables
3. Mapping between undocumented and official variables

Features:
- Uses official 27 employment variables from TM2 documentation
- Searches for aggregated and variant employment variables
- Maps undocumented variables back to official ones
- Comprehensive pattern matching and formula analysis
- Detailed reporting with aggregation analysis

Author: Comprehensive Analysis
Date: August 2025
"""

import os
import pandas as pd
import openpyxl
from pathlib import Path
import re
from collections import defaultdict, Counter
import warnings
import json
import argparse

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

class ComprehensiveEmploymentAnalyzer:
    def __init__(self, directory_path, tm2_directory=None, output_directory=None, verbose=False):
        """
        Initialize the comprehensive employment analyzer.
        
        Args:
            directory_path (str|Path): Path to the directory containing Excel coefficient files
            tm2_directory (str|Path, optional): Path to the TM2 root directory
            output_directory (str|Path, optional): Path to save output files (default: same as directory_path)
            verbose (bool): Enable verbose output during analysis
        """
        self.directory_path = Path(directory_path)
        self.tm2_directory = Path(tm2_directory) if tm2_directory else None
        self.output_directory = Path(output_directory) if output_directory else self.directory_path
        self.verbose = verbose
        
        # Ensure output directory exists
        self.output_directory.mkdir(parents=True, exist_ok=True)
        
        if self.verbose:
            print(f"Initialized analyzer:")
            print(f"  Model coefficient directory: {self.directory_path}")
            print(f"  TM2 root directory: {self.tm2_directory}")
            print(f"  Output directory: {self.output_directory}")
            print(f"  Verbose mode: {self.verbose}")
            print()
        
        # Official employment variable names from Bay Area Metro TM2 documentation
        self.official_employment_variables = [
            'ag',           # Agriculture employment
            'art_rec',      # Arts & recreational employment  
            'constr',       # Construction employment
            'eat',          # Eating out employment
            'ed_high',      # Higher education employment
            'ed_k12',       # K-12 education employment
            'ed_oth',       # Other education employment
            'fire',         # Financial, Insurance, real estate employment
            'gov',          # Government employment
            'health',       # Health employment
            'hotel',        # Hotel employment
            'info',         # Information employment
            'lease',        # Leasing employment
            'logis',        # Logistics employment
            'man_bio',      # Biological manufacturing employment
            'man_hvy',      # Heavy manufacturing employment
            'man_lgt',      # Light manufacturing employment
            'man_tech',     # Technology manufacturing employment
            'natres',       # Natural resources employment
            'prof',         # Professional employment
            'ret_loc',      # Local retail employment
            'ret_reg',      # Regional retail employment
            'serv_bus',     # Business services employment
            'serv_pers',    # Personal services employment
            'serv_soc',     # Social services employment
            'transp',       # Transportation employment
            'util'          # Utilities employment
        ]
        
        # Common aggregated variables that might be used
        self.aggregated_variables = [
            'emp_total',     # Total employment
            'empTotal',      # Alternative total employment
            'totemp',        # Another total employment variant
            'education',     # Aggregated education (ed_high + ed_k12 + ed_oth)
            'retail',        # Aggregated retail (ret_loc + ret_reg)
            'manufacturing', # Aggregated manufacturing (man_bio + man_hvy + man_lgt + man_tech)
            'services',      # Aggregated services (serv_bus + serv_pers + serv_soc)
            'empEduHealth',  # Education + Health aggregation
        ]
        
        # Undocumented variables found in legacy analysis scripts
        self.undocumented_variables = [
            # Construction variants (TM1 style)
            'emp_ag', 'emp_const_non_bldg_prod', 'emp_const_non_bldg_office', 
            'emp_const_bldg_prod', 'emp_const_bldg_office',
            
            # Manufacturing variants
            'emp_mfg_prod', 'emp_mfg_office',
            
            # Retail/Wholesale variants
            'emp_whsle_whs', 'emp_retail_loc', 'emp_retail_reg',
            
            # Professional services variants
            'emp_prof_bus_svcs', 'emp_prof_bus_svcs_bldg_maint',
            
            # Education variants
            'emp_pvt_ed_k12', 'emp_pvt_ed_post_k12_oth', 'emp_public_ed',
            
            # Other service variants
            'emp_health', 'emp_hotel', 'emp_restaurant_bar',
            'emp_personal_svcs_office', 'emp_personal_svcs_retail',
            'emp_amusement', 'emp_othr_svcs',
            
            # Government variants
            'emp_fed_non_mil', 'emp_fed_mil', 'emp_state_local_gov_blue', 'emp_state_local_gov_white',
            
            # Alternative patterns
            'agriculture', 'construction', 'manufacturing_food', 'manufacturing_nonfood',
            'wholesale', 'transport', 'information', 'finance', 'real_estate',
            'professional', 'management', 'administrative', 'arts', 'accommodation', 
            'food_service', 'other_services', 'military', 'utilities', 'office', 'industrial'
        ]
        
        # Combine all variables to search for
        self.all_variables = (self.official_employment_variables + 
                             self.aggregated_variables + 
                             self.undocumented_variables)
        
        # Generate variable patterns for comprehensive matching
        self.variable_patterns = []
        for var in self.all_variables:
            self.variable_patterns.extend([
                var,                    # exact match
                f'emp_{var}',          # with emp_ prefix
                f'{var}_emp',          # with _emp suffix
                f'{var}Emp',           # camelCase
                f'{var}Employment',    # with Employment suffix
                f'{var}Access',        # accessibility variables
                f'{var}Accessibility', # accessibility variables
                f'{var}Density',       # density variables
            ])
        
        # Remove duplicates
        self.variable_patterns = list(set(self.variable_patterns))
        
        # Results storage
        self.results = defaultdict(lambda: defaultdict(list))
        self.variable_counts = defaultdict(int)
        self.model_usage = defaultdict(set)
        
        # Variable mappings for undocumented variables
        self.variable_mappings = self._create_variable_mappings()
        
    def _create_variable_mappings(self):
        """Create mappings from undocumented variables to official ones."""
        return {
            # Construction variants
            'emp_const_non_bldg_prod': 'constr',
            'emp_const_non_bldg_office': 'constr', 
            'emp_const_bldg_prod': 'constr',
            'emp_const_bldg_office': 'constr',
            'construction': 'constr',
            
            # Manufacturing variants
            'emp_mfg_prod': ['man_lgt', 'man_hvy', 'man_tech', 'man_bio'],
            'emp_mfg_office': ['man_lgt', 'man_hvy', 'man_tech', 'man_bio'],
            'manufacturing_food': 'man_bio',
            'manufacturing_nonfood': ['man_lgt', 'man_hvy', 'man_tech'],
            
            # Retail/Wholesale variants
            'emp_whsle_whs': 'logis',
            'emp_retail_loc': 'ret_loc',
            'emp_retail_reg': 'ret_reg',
            'wholesale': 'logis',
            'retail': ['ret_loc', 'ret_reg'],
            
            # Professional services
            'emp_prof_bus_svcs': ['prof', 'serv_bus'],
            'emp_prof_bus_svcs_bldg_maint': ['prof', 'serv_bus'],
            'professional': 'prof',
            'management': 'prof',
            'administrative': 'serv_bus',
            
            # Education variants
            'emp_pvt_ed_k12': 'ed_k12',
            'emp_pvt_ed_post_k12_oth': ['ed_high', 'ed_oth'],
            'emp_public_ed': ['ed_k12', 'ed_high', 'ed_oth'],
            'education': ['ed_k12', 'ed_high', 'ed_oth'],
            'empEduHealth': ['ed_k12', 'ed_high', 'ed_oth', 'health'],
            
            # Health
            'emp_health': 'health',
            
            # Hotel/accommodation
            'emp_hotel': 'hotel',
            'accommodation': 'hotel',
            
            # Food service
            'emp_restaurant_bar': 'eat',
            'food_service': 'eat',
            
            # Personal services
            'emp_personal_svcs_office': 'serv_pers',
            'emp_personal_svcs_retail': 'serv_pers',
            
            # Entertainment
            'emp_amusement': 'art_rec',
            'arts': 'art_rec',
            
            # Other services
            'emp_othr_svcs': ['serv_bus', 'serv_pers', 'serv_soc'],
            'other_services': ['serv_bus', 'serv_pers', 'serv_soc'],
            
            # Government
            'emp_fed_non_mil': 'gov',
            'emp_fed_mil': 'gov',
            'emp_state_local_gov_blue': 'gov',
            'emp_state_local_gov_white': 'gov',
            'military': 'gov',
            
            # Agriculture
            'emp_ag': 'ag',
            'agriculture': 'ag',
            
            # Total employment
            'empTotal': 'Total of all employment categories',
            'emp_total': 'Total of all employment categories',
            'totemp': 'Total of all employment categories',
            
            # Alternative patterns
            'transport': 'transp',
            'information': 'info',
            'finance': 'fire',
            'real_estate': 'fire',
            'utilities': 'util',
            'office': ['prof', 'serv_bus', 'fire', 'info'],
            'industrial': ['man_lgt', 'man_hvy', 'man_tech', 'man_bio', 'util'],
        }
    
    def is_employment_variable(self, text):
        """
        Check if text is an employment variable name.
        
        Args:
            text (str): Text to check
            
        Returns:
            tuple: (is_employment_var, matched_variable, variable_type)
        """
        if not isinstance(text, str):
            text = str(text)
        
        text_clean = text.strip().lower()
        
        # Skip very short text that might be false positives
        if len(text_clean) < 2:
            return False, None, None
        
        # Check for exact matches first
        for var in self.variable_patterns:
            if text_clean == var.lower():
                base_var = self.get_base_variable(var)
                var_type = self.get_variable_type(base_var)
                return True, base_var, var_type
        
        # For short variables like 'ag', 'eat' - require stricter context
        short_vars = ['ag', 'eat', 'gov', 'util', 'info', 'fire', 'prof']
        
        # Check for variable references in formulas or text - STRICTER MATCHING
        import re
        for var in self.variable_patterns:
            var_lower = var.lower()
            
            # For short variables, require employment context nearby
            if any(short_var in var_lower for short_var in short_vars):
                # Look for employment context in the same text
                if not any(emp_word in text_clean for emp_word in ['emp', 'employment', 'employ']):
                    continue
            
            # Use word boundary regex for exact matches only
            # This pattern ensures the variable is surrounded by non-alphanumeric characters
            pattern = r'\b' + re.escape(var_lower) + r'\b'
            if re.search(pattern, text_clean):
                base_var = self.get_base_variable(var)
                var_type = self.get_variable_type(base_var)
                return True, base_var, var_type
            
            # Check for @variable references (Excel formulas) - exact match only
            if f'@{var_lower}' in text_clean:
                base_var = self.get_base_variable(var)
                var_type = self.get_variable_type(base_var)
                return True, base_var, var_type
        
        return False, None, None
    
    def get_base_variable(self, var_pattern):
        """Get the base variable name from a pattern match."""
        var_lower = var_pattern.lower()
        
        # Check if it's one of our base variables
        for base_var in self.all_variables:
            if var_lower == base_var.lower():
                return base_var
            if var_lower == f'emp_{base_var}'.lower():
                return base_var
            if var_lower == f'{base_var}_emp'.lower():
                return base_var
            if var_lower == f'{base_var}emp'.lower():
                return base_var
            if var_lower == f'{base_var}employment'.lower():
                return base_var
            if var_lower == f'{base_var}access'.lower():
                return base_var
            if var_lower == f'{base_var}accessibility'.lower():
                return base_var
            if var_lower == f'{base_var}density'.lower():
                return base_var
        
        return var_pattern  # fallback
    
    def get_variable_type(self, variable):
        """Determine the type of employment variable."""
        if variable in self.official_employment_variables:
            return 'official'
        elif variable in self.aggregated_variables:
            return 'aggregated'
        elif variable in self.undocumented_variables:
            return 'undocumented'
        else:
            return 'unknown'
    
    def extract_formula_variables(self, text):
        """Extract employment variables from Excel formulas."""
        if not isinstance(text, str):
            return []
        
        found_vars = []
        text_lower = text.lower()
        
        import re
        for var in self.variable_patterns:
            var_lower = var.lower()
            
            # Use word boundary regex for more precise matching
            pattern = r'\b' + re.escape(var_lower) + r'\b'
            if re.search(pattern, text_lower):
                base_var = self.get_base_variable(var)
                found_vars.append(base_var)
        
        return list(set(found_vars))  # Remove duplicates
    
    def analyze_excel_file(self, file_path):
        """Analyze a single Excel file for employment variables."""
        if self.verbose:
            print(f"Analyzing: {file_path.name}")
        model_name = file_path.name.replace('.xls', '').replace('.xlsx', '')
        
        try:
            # Try reading with pandas first
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names
            
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    self.analyze_dataframe(df, file_path.name, sheet_name, model_name)
                except Exception as e:
                    if self.verbose:
                        print(f"  Warning: Could not read sheet '{sheet_name}': {e}")
                    
        except Exception as e:
            # Try openpyxl if pandas fails
            try:
                wb = openpyxl.load_workbook(file_path, data_only=False)  # Keep formulas
                for sheet_name in wb.sheetnames:
                    try:
                        ws = wb[sheet_name]
                        self.analyze_worksheet(ws, file_path.name, sheet_name, model_name)
                    except Exception as e:
                        if self.verbose:
                            print(f"  Warning: Could not analyze sheet '{sheet_name}': {e}")
                wb.close()
            except Exception as e:
                if self.verbose:
                    print(f"  Error: Could not open {file_path.name}: {e}")
                else:
                    print(f"Skipping {file_path.name}: {e}")
    
    def has_employment_context(self, df):
        """Check if DataFrame contains employment-related context."""
        # Convert entire dataframe to string and check for employment keywords
        try:
            df_str = df.astype(str).values.flatten()
            text_content = ' '.join(df_str).lower()
            
            # Look for employment-related keywords
            employment_keywords = ['emp', 'employment', 'employ', 'job', 'worker', 'sector', 'industry']
            return any(keyword in text_content for keyword in employment_keywords)
        except:
            return False
    
    def analyze_dataframe(self, df, file_name, sheet_name, model_name):
        """Analyze a pandas DataFrame for employment variables."""
        found_vars = set()
        
        # Quick check: Skip sheets without employment context
        if not self.has_employment_context(df):
            if self.verbose:
                print(f"  Sheet '{sheet_name}': Skipped (no employment context)")
            return
        
        # Check all cells for employment variables (reduced scope for speed)
        max_rows = min(50, len(df))  # Reduced from 200 to 50
        max_cols = min(20, len(df.columns))  # Reduced from 30 to 20
        
        for row_idx in range(max_rows):
            for col_idx in range(max_cols):
                try:
                    cell_value = df.iloc[row_idx, col_idx]
                    if pd.notna(cell_value):
                        cell_str = str(cell_value).strip()
                        
                        # Skip very long strings (descriptions)
                        if len(cell_str) > 100:
                            continue
                        
                        # Check if it's an employment variable
                        is_emp_var, variable, var_type = self.is_employment_variable(cell_str)
                        if is_emp_var:
                            found_vars.add(variable)
                            self.variable_counts[variable] += 1
                            self.model_usage[model_name].add(variable)
                            self.results[file_name][sheet_name].append({
                                'location': f'Row {row_idx + 1}, Col {col_idx + 1}',
                                'content': cell_str,
                                'variable': variable,
                                'variable_type': var_type,
                                'type': 'direct_variable'
                            })
                        
                        # Check for formula variables
                        formula_vars = self.extract_formula_variables(cell_str)
                        for var in formula_vars:
                            found_vars.add(var)
                            self.variable_counts[var] += 1
                            self.model_usage[model_name].add(var)
                            var_type = self.get_variable_type(var)
                            self.results[file_name][sheet_name].append({
                                'location': f'Row {row_idx + 1}, Col {col_idx + 1}',
                                'content': cell_str[:100],
                                'variable': var,
                                'variable_type': var_type,
                                'type': 'formula_variable'
                            })
                            
                except Exception:
                    continue
        
        if found_vars:
            if self.verbose:
                print(f"  Sheet '{sheet_name}': Found {len(found_vars)} employment variables: {sorted(found_vars)}")
            else:
                print(f"  {file_name}[{sheet_name}]: {len(found_vars)} variables")
    
    def has_employment_context_worksheet(self, ws):
        """Check if worksheet contains employment-related context."""
        try:
            # Sample a few cells to check for employment context
            max_row = min(ws.max_row or 50, 50)
            max_col = min(ws.max_column or 20, 20)
            
            sample_text = ""
            for row in range(1, max_row + 1, 5):  # Sample every 5th row
                for col in range(1, max_col + 1, 5):  # Sample every 5th column
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        sample_text += str(cell.value).lower() + " "
            
            # Look for employment-related keywords
            employment_keywords = ['emp', 'employment', 'employ', 'job', 'worker', 'sector', 'industry']
            return any(keyword in sample_text for keyword in employment_keywords)
        except:
            return False
    
    def analyze_worksheet(self, ws, file_name, sheet_name, model_name):
        """Analyze an openpyxl worksheet for employment variables."""
        found_vars = set()
        
        # Quick check: Skip sheets without employment context
        if not self.has_employment_context_worksheet(ws):
            if self.verbose:
                print(f"  Sheet '{sheet_name}': Skipped (no employment context)")
            return
        
        max_row = min(ws.max_row or 50, 50)  # Reduced from 200 to 50
        max_col = min(ws.max_column or 20, 20)  # Reduced from 30 to 20
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    
                    # Check cell value
                    if cell.value:
                        cell_str = str(cell.value).strip()
                        
                        # Skip very long strings
                        if len(cell_str) > 100:
                            continue
                        
                        is_emp_var, variable, var_type = self.is_employment_variable(cell_str)
                        if is_emp_var:
                            found_vars.add(variable)
                            self.variable_counts[variable] += 1
                            self.model_usage[model_name].add(variable)
                            self.results[file_name][sheet_name].append({
                                'location': f'Row {row}, Col {col}',
                                'content': cell_str,
                                'variable': variable,
                                'variable_type': var_type,
                                'type': 'direct_variable'
                            })
                    
                    # Check cell formula if it exists
                    if hasattr(cell, 'formula') and cell.formula:
                        formula_vars = self.extract_formula_variables(cell.formula)
                        for var in formula_vars:
                            found_vars.add(var)
                            self.variable_counts[var] += 1
                            self.model_usage[model_name].add(var)
                            var_type = self.get_variable_type(var)
                            self.results[file_name][sheet_name].append({
                                'location': f'Row {row}, Col {col}',
                                'content': cell.formula[:100],
                                'variable': var,
                                'variable_type': var_type,
                                'type': 'formula_variable'
                            })
                    
                except Exception:
                    continue
        
        if found_vars:
            if self.verbose:
                print(f"  Sheet '{sheet_name}': Found {len(found_vars)} employment variables: {sorted(found_vars)}")
            else:
                print(f"  {file_name}[{sheet_name}]: {len(found_vars)} variables")
    
    def run_analysis(self):
        """Run the complete analysis on all Excel files in the directory."""
        print(f"Starting comprehensive employment variable analysis in: {self.directory_path}")
        print("=" * 90)
        print(f"Looking for {len(self.official_employment_variables)} official employment variables:")
        print(f"  {', '.join(self.official_employment_variables)}")
        print(f"Plus {len(self.aggregated_variables)} aggregated variables and {len(self.undocumented_variables)} undocumented variables")
        print("=" * 90)
        
        # Find all Excel files
        excel_files = []
        for pattern in ['*.xls', '*.xlsx']:
            excel_files.extend(self.directory_path.glob(pattern))
        
        if not excel_files:
            print("No Excel files found in the directory.")
            return
        
        print(f"\nFound {len(excel_files)} Excel files to analyze.\n")
        
        # Analyze each file
        for file_path in sorted(excel_files):
            # Skip temporary files
            if file_path.name.startswith('._') or file_path.name.endswith('.tmp'):
                continue
            self.analyze_excel_file(file_path)
        
        # Generate comprehensive report
        self.generate_comprehensive_report()
    
    def generate_comprehensive_report(self):
        """Generate a comprehensive report combining all analysis types."""
        print("\n" + "=" * 90)
        print("COMPREHENSIVE EMPLOYMENT VARIABLE ANALYSIS REPORT")
        print("=" * 90)
        print("This report shows how employment variables are used across TM2 model files.")
        print("'Occurrences' = total times variable found in Excel cells (headers, formulas, values, etc.)")
        print("'Models' = number of different Excel files containing the variable")
        print("=" * 90)
        
        if not self.variable_counts:
            print("No employment variables found in any Excel files.")
            return
        
        # Categorize found variables
        official_found = [v for v in self.variable_counts.keys() if v in self.official_employment_variables]
        aggregated_found = [v for v in self.variable_counts.keys() if v in self.aggregated_variables]
        undocumented_found = [v for v in self.variable_counts.keys() if v in self.undocumented_variables]
        unknown_found = [v for v in self.variable_counts.keys() if v not in self.all_variables]
        
        # Calculate statistics
        total_vars_found = len(self.variable_counts)
        total_files_analyzed = len([f for f in self.results if self.results[f]])
        
        print("SUMMARY STATISTICS:")
        print("-" * 20)
        print(f"Total official employment variables:     {len(self.official_employment_variables):3d}")
        print(f"Official variables found in models:     {len(official_found):3d}")
        print(f"Aggregated variables found:              {len(aggregated_found):3d}")
        print(f"Undocumented variables found:            {len(undocumented_found):3d}")
        print(f"Unknown variables found:                 {len(unknown_found):3d}")
        print(f"Total variables found:                   {total_vars_found:3d}")
        print(f"Models with employment variables:        {len(self.model_usage):3d}")
        print(f"Files analyzed with findings:            {total_files_analyzed:3d}")
        print()
        
        # Variable usage summary by type
        print("VARIABLE USAGE BY TYPE:")
        print("-" * 25)
        print("NOTE: 'Occurrences' = total times variable found across all Excel cells")
        print("      'Models' = number of different Excel files containing the variable")
        print()
        
        if official_found:
            print(f"OFFICIAL VARIABLES ({len(official_found)} found):")
            print("These are the 27 official TM2 employment variables from Bay Area Metro documentation")
            sorted_official = sorted([(v, self.variable_counts[v]) for v in official_found], 
                                   key=lambda x: x[1], reverse=True)
            for variable, count in sorted_official:
                models_using = len([m for m, vars in self.model_usage.items() if variable in vars])
                print(f"  {variable:15} {count:4d} occurrences in {models_using:2d} models")
        
        if aggregated_found:
            print(f"\nAGGREGATED VARIABLES ({len(aggregated_found)} found):")
            print("These are combinations of official variables (e.g., 'retail' = ret_loc + ret_reg)")
            sorted_aggregated = sorted([(v, self.variable_counts[v]) for v in aggregated_found], 
                                     key=lambda x: x[1], reverse=True)
            for variable, count in sorted_aggregated:
                models_using = len([m for m, vars in self.model_usage.items() if variable in vars])
                mapping = self.variable_mappings.get(variable, 'Unknown mapping')
                if isinstance(mapping, list):
                    mapping = ' + '.join(mapping)
                print(f"  {variable:15} {count:4d} occurrences in {models_using:2d} models → {mapping}")
        
        if undocumented_found:
            print(f"\nUNDOCUMENTED VARIABLES ({len(undocumented_found)} found):")
            print("These are legacy or alternative names that map to official variables")
            sorted_undocumented = sorted([(v, self.variable_counts[v]) for v in undocumented_found], 
                                       key=lambda x: x[1], reverse=True)
            for variable, count in sorted_undocumented:
                models_using = len([m for m, vars in self.model_usage.items() if variable in vars])
                mapping = self.variable_mappings.get(variable, 'Unknown mapping')
                if isinstance(mapping, list):
                    mapping = ' + '.join(mapping)
                print(f"  {variable:20} {count:4d} occurrences in {models_using:2d} models → {mapping}")
        
        if unknown_found:
            print(f"\nUNKNOWN VARIABLES ({len(unknown_found)} found):")
            sorted_unknown = sorted([(v, self.variable_counts[v]) for v in unknown_found], 
                                  key=lambda x: x[1], reverse=True)
            for variable, count in sorted_unknown:
                models_using = len([m for m, vars in self.model_usage.items() if variable in vars])
                print(f"  {variable:20} {count:4d} occurrences in {models_using:2d} models")
        
        # Official variables not found
        print(f"\nOFFICIAL VARIABLES NOT FOUND ({len(self.official_employment_variables) - len(official_found)}):")
        print("-" * 35)
        unused_official = set(self.official_employment_variables) - set(official_found)
        if unused_official:
            for var in sorted(unused_official):
                print(f"  - {var}")
        else:
            print("  All official variables are used!")
        
        # Model-by-model breakdown
        print(f"\nMODEL-BY-MODEL BREAKDOWN:")
        print("-" * 25)
        for model, variables in sorted(self.model_usage.items()):
            official_vars = [v for v in variables if v in self.official_employment_variables]
            aggregated_vars = [v for v in variables if v in self.aggregated_variables]
            undocumented_vars = [v for v in variables if v in self.undocumented_variables]
            
            print(f"\n{model} ({len(variables)} total variables):")
            if official_vars:
                print(f"  Official ({len(official_vars)}): {', '.join(sorted(official_vars))}")
            if aggregated_vars:
                print(f"  Aggregated ({len(aggregated_vars)}): {', '.join(sorted(aggregated_vars))}")
            if undocumented_vars:
                print(f"  Undocumented ({len(undocumented_vars)}): {', '.join(sorted(undocumented_vars))}")
        
        # Aggregation analysis
        self.analyze_aggregation_patterns()
        
        # Save comprehensive reports
        self.save_comprehensive_reports()
    
    def analyze_aggregation_patterns(self):
        """Analyze aggregation patterns in the models."""
        print(f"\n" + "=" * 60)
        print("AGGREGATION PATTERN ANALYSIS")
        print("=" * 60)
        
        # Check which models use aggregated vs detailed variables
        models_with_official = set()
        models_with_aggregated = set()
        models_with_undocumented = set()
        
        for model, variables in self.model_usage.items():
            has_official = any(v in self.official_employment_variables for v in variables)
            has_aggregated = any(v in self.aggregated_variables for v in variables)
            has_undocumented = any(v in self.undocumented_variables for v in variables)
            
            if has_official:
                models_with_official.add(model)
            if has_aggregated:
                models_with_aggregated.add(model)
            if has_undocumented:
                models_with_undocumented.add(model)
        
        print(f"Models using official variables:      {len(models_with_official)}")
        print(f"Models using aggregated variables:    {len(models_with_aggregated)}")
        print(f"Models using undocumented variables:  {len(models_with_undocumented)}")
        print(f"Models using official + aggregated:   {len(models_with_official & models_with_aggregated)}")
        print(f"Models using all three types:         {len(models_with_official & models_with_aggregated & models_with_undocumented)}")
        
        # Specific aggregation examples
        retail_detailed = [m for m, vars in self.model_usage.items() if any(v in ['ret_loc', 'ret_reg'] for v in vars)]
        retail_aggregated = [m for m, vars in self.model_usage.items() if 'retail' in vars]
        
        edu_detailed = [m for m, vars in self.model_usage.items() if any(v in ['ed_high', 'ed_k12', 'ed_oth'] for v in vars)]
        edu_aggregated = [m for m, vars in self.model_usage.items() if 'education' in vars or 'empEduHealth' in vars]
        
        print(f"\nSpecific Aggregation Examples:")
        print(f"  Retail - detailed (ret_loc/ret_reg):     {len(retail_detailed)} models")
        print(f"  Retail - aggregated ('retail'):          {len(retail_aggregated)} models")
        print(f"  Education - detailed (ed_*):             {len(edu_detailed)} models")
        print(f"  Education - aggregated:                  {len(edu_aggregated)} models")
    
    def save_comprehensive_reports(self):
        """Save comprehensive reports to files."""
        # Save detailed analysis as JSON
        results_data = {
            'metadata': {
                'generated_on': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                'model_directory_analyzed': str(self.directory_path),
                'tm2_directory': str(self.tm2_directory) if self.tm2_directory else None,
                'output_directory': str(self.output_directory),
                'total_files_analyzed': len([f for f in self.results if self.results[f]]),
                'total_variables_found': len(self.variable_counts)
            },
            'variable_definitions': {
                'official_employment_variables': self.official_employment_variables,
                'aggregated_variables': self.aggregated_variables,
                'undocumented_variables': self.undocumented_variables
            },
            'variable_mappings': self.variable_mappings,
            'analysis_results': {
                'variable_counts': dict(self.variable_counts),
                'model_usage': {model: list(vars) for model, vars in self.model_usage.items()},
                'detailed_findings': dict(self.results)
            }
        }
        
        json_path = self.output_directory / "employment_analysis_comprehensive.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(results_data, f, indent=2, default=str)
        
        # Save detailed text report
        report_path = self.output_directory / "employment_analysis_comprehensive.txt"
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("COMPREHENSIVE EMPLOYMENT VARIABLE ANALYSIS\n")
            f.write("=" * 50 + "\n")
            f.write(f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Model directory analyzed: {self.directory_path}\n")
            if self.tm2_directory:
                f.write(f"TM2 root directory: {self.tm2_directory}\n")
            f.write(f"Output directory: {self.output_directory}\n\n")
            
            # Add explanation section
            f.write("HOW TO READ THIS REPORT:\n")
            f.write("-" * 25 + "\n")
            f.write("OCCURRENCES: Total number of times a variable was found across all Excel cells.\n")
            f.write("            This includes: cell values, formula references, column headers,\n")
            f.write("            coefficient tables, and any other text mentioning the variable.\n\n")
            f.write("MODELS:     Number of different Excel model files containing the variable.\n")
            f.write("            Each Excel file represents a different travel model component\n")
            f.write("            (e.g., TourModeChoice.xls, AutoOwnership.xls, etc.).\n\n")
            f.write("VARIABLE TYPES:\n")
            f.write("  - Official:      The 27 employment variables from TM2 documentation\n")
            f.write("  - Aggregated:    Combinations of official variables (e.g., 'retail' = ret_loc + ret_reg)\n")
            f.write("  - Undocumented:  Legacy or alternative names that map to official variables\n")
            f.write("  - Unknown:       Variables not in any of the above categories\n\n")
            f.write("EXAMPLE: 'ag    183 occurrences in 14 models (official)' means:\n")
            f.write("         The agriculture employment variable 'ag' was found 183 times\n")
            f.write("         across 14 different Excel model files, and it's an official TM2 variable.\n\n")
            f.write("=" * 70 + "\n\n")
            
            # Variable definitions
            f.write("OFFICIAL EMPLOYMENT VARIABLES (27):\n")
            f.write("-" * 35 + "\n")
            for i, var in enumerate(self.official_employment_variables, 1):
                f.write(f"{i:2d}. {var}\n")
            
            f.write("\nAGGREGATED VARIABLES:\n")
            f.write("-" * 20 + "\n")
            for var in self.aggregated_variables:
                f.write(f"  - {var}\n")
            
            # Variable usage summary
            f.write("\nVARIABLE USAGE SUMMARY:\n")
            f.write("-" * 25 + "\n")
            f.write("NOTE: 'Occurrences' = total times variable found across all Excel cells\n")
            f.write("      'Models' = number of different Excel files containing the variable\n")
            f.write("      Each occurrence could be: cell values, formula references, headers, etc.\n\n")
            sorted_vars = sorted(self.variable_counts.items(), key=lambda x: x[1], reverse=True)
            for variable, count in sorted_vars:
                models_using = len([m for m, vars in self.model_usage.items() if variable in vars])
                var_type = self.get_variable_type(variable)
                f.write(f"{variable:20} {count:4d} occurrences in {models_using:2d} models ({var_type})\n")
            
            # Variable mappings
            f.write("\nUNDOCUMENTED → OFFICIAL VARIABLE MAPPINGS:\n")
            f.write("-" * 45 + "\n")
            for undoc_var, official_mapping in sorted(self.variable_mappings.items()):
                if isinstance(official_mapping, list):
                    official_mapping = ' + '.join(official_mapping)
                f.write(f"{undoc_var:25} → {official_mapping}\n")
            
            # Model details
            f.write("\nMODEL DETAILS:\n")
            f.write("-" * 15 + "\n")
            for model, variables in sorted(self.model_usage.items()):
                f.write(f"\n{model} ({len(variables)} variables):\n")
                for var in sorted(variables):
                    var_type = self.get_variable_type(var)
                    count = self.variable_counts[var]
                    f.write(f"  - {var} ({count} occurrences, {var_type})\n")
        
        print(f"\nComprehensive reports saved:")
        print(f"  JSON: {json_path}")
        print(f"  Text: {report_path}")


def main():
    """Main function to run the comprehensive employment variable analysis."""
    
    # Set up command line argument parsing
    parser = argparse.ArgumentParser(
        description='Comprehensive Employment Variable Analysis for Travel Model Two',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default model coefficient directory
  python employment_analysis_comprehensive.py
  
  # Specify custom model coefficient directory
  python employment_analysis_comprehensive.py --model-dir "C:/path/to/model/coefficients"
  
  # Specify both model directories
  python employment_analysis_comprehensive.py \\
    --model-dir "C:/GitHub/travel-model-two/model-files/model" \\
    --tm2-dir "C:/GitHub/travel-model-two"
        """
    )
    
    # Default paths
    default_model_dir = r"C:\GitHub\travel-model-two\model-files\model"
    default_tm2_dir = r"C:\GitHub\travel-model-two"
    
    parser.add_argument(
        '--model-dir', '--model-directory',
        default=default_model_dir,
        help=f'Path to the directory containing model coefficient Excel files (default: {default_model_dir})'
    )
    
    parser.add_argument(
        '--tm2-dir', '--tm2-directory',
        default=default_tm2_dir,
        help=f'Path to the Travel Model Two root directory (default: {default_tm2_dir})'
    )
    
    parser.add_argument(
        '--output-dir', '--output-directory',
        default=None,
        help='Path to save output reports (default: same as model-dir)'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose output during analysis'
    )
    
    # Parse arguments
    args = parser.parse_args()
    
    # Validate directories
    model_dir = Path(args.model_dir)
    tm2_dir = Path(args.tm2_dir)
    output_dir = Path(args.output_dir) if args.output_dir else model_dir
    
    if not model_dir.exists():
        print(f"Error: Model directory does not exist: {model_dir}")
        print("Please check the path or use --model-dir to specify the correct location.")
        return 1
    
    if not tm2_dir.exists():
        print(f"Error: TM2 directory does not exist: {tm2_dir}")
        print("Please check the path or use --tm2-dir to specify the correct location.")
        return 1
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print("=" * 80)
    print("COMPREHENSIVE EMPLOYMENT VARIABLE ANALYSIS")
    print("=" * 80)
    print(f"Model coefficient directory: {model_dir}")
    print(f"TM2 root directory:          {tm2_dir}")
    print(f"Output directory:            {output_dir}")
    print(f"Verbose mode:                {'Enabled' if args.verbose else 'Disabled'}")
    print("=" * 80)
    
    # Run the analysis
    analyzer = ComprehensiveEmploymentAnalyzer(
        directory_path=model_dir,
        tm2_directory=tm2_dir,
        output_directory=output_dir,
        verbose=args.verbose
    )
    analyzer.run_analysis()
    
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())

# Example usage:
"""
# Run with default directories
python employment_analysis_comprehensive.py

# Run with custom model directory
python employment_analysis_comprehensive.py --model-dir "C:/custom/model/path"

# Run with verbose output
python employment_analysis_comprehensive.py --verbose

# Run with custom output directory
python employment_analysis_comprehensive.py --output-dir "C:/analysis/results"

# Run with all custom paths
python employment_analysis_comprehensive.py \
    --model-dir "C:/GitHub/travel-model-two/model-files/model" \
    --tm2-dir "C:/GitHub/travel-model-two" \
    --output-dir "C:/analysis/results" \
    --verbose
"""
