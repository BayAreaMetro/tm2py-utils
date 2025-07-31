#!/usr/bin/env python3
"""
Employment Variable Summary Script

This script analyzes Excel files in the travel model directory to find and summarize
employment-related data across all sheets and files.

Author: Generated for TM2 Analysis
Date: July 2025
"""

import os
import pandas as pd
import openpyxl
from pathlib import Path
import re
from collections import defaultdict
import warnings

# Suppress openpyxl warnings for better output readability
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

class EmploymentVariableSummarizer:
    def __init__(self, directory_path):
        """
        Initialize the employment variable summarizer.
        
        Args:
            directory_path (str): Path to the directory containing Excel files
        """
        self.directory_path = Path(directory_path)
        self.employment_keywords = [
            # Common employment-related terms
            'emp', 'employment', 'employ', 'job', 'jobs', 'worker', 'workers',
            'work', 'workplace', 'labor', 'labour', 'occupation', 'industry',
            'sector', 'business', 'commercial', 'office', 'retail', 'service',
            'manufacturing', 'agriculture', 'construction', 'transport',
            'government', 'edu', 'education', 'health', 'medical', 'finance',
            'professional', 'technical', 'administrative', 'management',
            # Travel model specific employment terms
            'totemp', 'retl', 'food', 'svc', 'mfg', 'wsle', 'fin', 'prof',
            'mgmt', 'admin', 'edu', 'hlth', 'arts', 'other', 'agrcon',
            # Activity-based model employment categories
            'non_work', 'at_work', 'mandatory', 'home_based_work',
            'work_based', 'employer', 'employee'
        ]
        
        # Define the 27 employment categories typically used in travel models
        self.employment_categories = {
            'agriculture': ['agr', 'agriculture', 'agrcon', 'farm'],
            'mining': ['min', 'mining', 'extract'],
            'construction': ['con', 'construction', 'build'],
            'manufacturing_food': ['mfg_food', 'food_mfg', 'food'],
            'manufacturing_nonfood': ['mfg', 'manufacturing', 'mfg_nonfood'],
            'wholesale': ['wsle', 'wholesale', 'whsl'],
            'retail': ['retl', 'retail', 'shop'],
            'transport': ['trans', 'transport', 'transportation', 'ship'],
            'information': ['info', 'information', 'telecom', 'media'],
            'finance': ['fin', 'finance', 'bank', 'insurance'],
            'real_estate': ['real', 'estate', 'property'],
            'professional': ['prof', 'professional', 'legal', 'accounting'],
            'management': ['mgmt', 'management', 'admin'],
            'administrative': ['admin', 'administrative', 'support'],
            'education': ['edu', 'education', 'school', 'university'],
            'health': ['hlth', 'health', 'medical', 'hospital'],
            'arts': ['arts', 'entertainment', 'recreation'],
            'accommodation': ['hotel', 'accommodation', 'lodging'],
            'food_service': ['restaurant', 'food_svc', 'dining'],
            'other_services': ['svc', 'service', 'personal'],
            'government': ['gov', 'government', 'public'],
            'military': ['military', 'defense'],
            'utilities': ['util', 'utilities', 'power', 'water'],
            'total_employment': ['totemp', 'total_emp', 'emp_total'],
            'office': ['office', 'commercial'],
            'industrial': ['industrial', 'warehouse'],
            'other': ['other', 'misc']
        }
        
        self.results = defaultdict(list)
        self.category_usage = defaultdict(list)  # Track which models use which categories
        self.coefficient_patterns = defaultdict(list)  # Track coefficient values for categories
        
    def is_employment_related(self, text):
        """
        Check if a text string contains employment-related keywords.
        
        Args:
            text (str): Text to check
            
        Returns:
            bool: True if employment-related, False otherwise
        """
        if not isinstance(text, str):
            text = str(text)
        
        text_lower = text.lower()
        return any(keyword in text_lower for keyword in self.employment_keywords)
    
    def categorize_employment_variable(self, text):
        """
        Categorize an employment variable into one of the 27 standard categories.
        
        Args:
            text (str): Text to categorize
            
        Returns:
            list: List of matching employment categories
        """
        if not isinstance(text, str):
            text = str(text)
            
        text_lower = text.lower()
        found_categories = []
        
        for category, keywords in self.employment_categories.items():
            if any(keyword in text_lower for keyword in keywords):
                found_categories.append(category)
        
        return found_categories
    
    def extract_coefficient_value(self, text):
        """
        Extract numerical coefficient values from text.
        
        Args:
            text (str): Text to search for coefficients
            
        Returns:
            list: List of numerical values found
        """
        import re
        if not isinstance(text, str):
            text = str(text)
        
        # Look for decimal numbers (positive or negative)
        pattern = r'-?\d+\.?\d*'
        matches = re.findall(pattern, text)
        return [float(match) for match in matches if '.' in match or len(match) > 2]
    def extract_employment_keywords(self, text):
        """
        Extract specific employment keywords found in text.
        
        Args:
            text (str): Text to search
            
        Returns:
            list: List of found keywords
        """
        if not isinstance(text, str):
            text = str(text)
            
        text_lower = text.lower()
        found_keywords = []
        for keyword in self.employment_keywords:
            if keyword in text_lower:
                found_keywords.append(keyword)
        return found_keywords
    
    def analyze_excel_file(self, file_path):
        """
        Analyze a single Excel file for employment-related data.
        
        Args:
            file_path (Path): Path to the Excel file
        """
        print(f"Analyzing: {file_path.name}")
        
        try:
            # Try reading with pandas first (for .xls files)
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names
            
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    self.analyze_dataframe(df, file_path.name, sheet_name)
                except Exception as e:
                    print(f"  Warning: Could not read sheet '{sheet_name}' in {file_path.name}: {e}")
                    
        except Exception as e:
            # If pandas fails, try openpyxl
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    try:
                        ws = wb[sheet_name]
                        self.analyze_worksheet(ws, file_path.name, sheet_name)
                    except Exception as e:
                        print(f"  Warning: Could not analyze sheet '{sheet_name}' in {file_path.name}: {e}")
                wb.close()
            except Exception as e:
                print(f"  Error: Could not open {file_path.name}: {e}")
    
    def analyze_dataframe(self, df, file_name, sheet_name):
        """
        Analyze a pandas DataFrame for employment-related content.
        
        Args:
            df (pd.DataFrame): DataFrame to analyze
            file_name (str): Name of the source file
            sheet_name (str): Name of the sheet
        """
        employment_findings = []
        
        # Check column names
        for col_idx, col_name in enumerate(df.columns):
            if self.is_employment_related(col_name):
                keywords = self.extract_employment_keywords(col_name)
                categories = self.categorize_employment_variable(col_name)
                coefficients = self.extract_coefficient_value(col_name)
                
                employment_findings.append({
                    'location': f'Column {col_idx} header',
                    'content': str(col_name),
                    'keywords': keywords,
                    'categories': categories,
                    'coefficients': coefficients,
                    'type': 'column_header'
                })
                
                # Track category usage by model
                for category in categories:
                    self.category_usage[category].append({
                        'file': file_name,
                        'sheet': sheet_name,
                        'location': f'Column {col_idx} header',
                        'content': str(col_name)
                    })
        
        # Check cell values (sample first 100 rows to avoid performance issues)
        max_rows = min(100, len(df))
        for row_idx in range(max_rows):
            for col_idx in range(len(df.columns)):
                try:
                    cell_value = df.iloc[row_idx, col_idx]
                    if pd.notna(cell_value) and self.is_employment_related(cell_value):
                        keywords = self.extract_employment_keywords(cell_value)
                        categories = self.categorize_employment_variable(cell_value)
                        coefficients = self.extract_coefficient_value(cell_value)
                        
                        employment_findings.append({
                            'location': f'Row {row_idx + 1}, Col {col_idx + 1}',
                            'content': str(cell_value)[:100],  # Truncate long content
                            'keywords': keywords,
                            'categories': categories,
                            'coefficients': coefficients,
                            'type': 'cell_value'
                        })
                        
                        # Track category usage by model
                        for category in categories:
                            self.category_usage[category].append({
                                'file': file_name,
                                'sheet': sheet_name,
                                'location': f'Row {row_idx + 1}, Col {col_idx + 1}',
                                'content': str(cell_value)[:100]
                            })
                            
                        # Track coefficient patterns
                        for coeff in coefficients:
                            for category in categories:
                                self.coefficient_patterns[category].append({
                                    'file': file_name,
                                    'sheet': sheet_name,
                                    'value': coeff,
                                    'context': str(cell_value)[:100]
                                })
                                
                except Exception:
                    continue
        
        if employment_findings:
            self.results[file_name].append({
                'sheet_name': sheet_name,
                'findings': employment_findings
            })
    
    def analyze_worksheet(self, ws, file_name, sheet_name):
        """
        Analyze an openpyxl worksheet for employment-related content.
        
        Args:
            ws: openpyxl worksheet object
            file_name (str): Name of the source file
            sheet_name (str): Name of the sheet
        """
        employment_findings = []
        
        # Limit analysis to reasonable range to avoid performance issues
        max_row = min(ws.max_row, 100) if ws.max_row else 100
        max_col = min(ws.max_column, 50) if ws.max_column else 50
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and self.is_employment_related(cell.value):
                        keywords = self.extract_employment_keywords(cell.value)
                        categories = self.categorize_employment_variable(cell.value)
                        coefficients = self.extract_coefficient_value(cell.value)
                        
                        employment_findings.append({
                            'location': f'Row {row}, Col {col}',
                            'content': str(cell.value)[:100],  # Truncate long content
                            'keywords': keywords,
                            'categories': categories,
                            'coefficients': coefficients,
                            'type': 'cell_value'
                        })
                        
                        # Track category usage by model
                        for category in categories:
                            self.category_usage[category].append({
                                'file': file_name,
                                'sheet': sheet_name,
                                'location': f'Row {row}, Col {col}',
                                'content': str(cell.value)[:100]
                            })
                            
                        # Track coefficient patterns
                        for coeff in coefficients:
                            for category in categories:
                                self.coefficient_patterns[category].append({
                                    'file': file_name,
                                    'sheet': sheet_name,
                                    'value': coeff,
                                    'context': str(cell.value)[:100]
                                })
                                
                except Exception:
                    continue
        
        if employment_findings:
            self.results[file_name].append({
                'sheet_name': sheet_name,
                'findings': employment_findings
            })
    
    def run_analysis(self):
        """
        Run the complete analysis on all Excel files in the directory.
        """
        print(f"Starting employment variable analysis in: {self.directory_path}")
        print("=" * 60)
        
        # Find all Excel files
        excel_files = []
        for pattern in ['*.xls', '*.xlsx']:
            excel_files.extend(self.directory_path.glob(pattern))
        
        if not excel_files:
            print("No Excel files found in the directory.")
            return
        
        print(f"Found {len(excel_files)} Excel files to analyze.")
        print()
        
        # Analyze each file
        for file_path in sorted(excel_files):
            # Skip temporary files
            if file_path.name.startswith('._') and file_path.name.endswith('.tmp'):
                continue
            self.analyze_excel_file(file_path)
        
        # Generate summary report
        self.generate_report()
        
        # Generate advanced category analysis
        self.generate_category_analysis()
    
    def analyze_coefficient_similarity(self):
        """
        Analyze coefficient patterns to identify categories that might be aggregated.
        
        Returns:
            dict: Analysis of coefficient similarities
        """
        import numpy as np
        
        similarity_analysis = {}
        
        for category, coeffs in self.coefficient_patterns.items():
            if len(coeffs) > 1:
                values = [c['value'] for c in coeffs]
                similarity_analysis[category] = {
                    'count': len(values),
                    'mean': np.mean(values),
                    'std': np.std(values),
                    'min': np.min(values),
                    'max': np.max(values),
                    'coefficient_of_variation': np.std(values) / np.mean(values) if np.mean(values) != 0 else float('inf'),
                    'values': values
                }
        
        return similarity_analysis
    
    def identify_aggregation_candidates(self, similarity_threshold=0.2):
        """
        Identify employment categories that could potentially be aggregated.
        
        Args:
            similarity_threshold (float): Threshold for coefficient variation (lower = more similar)
            
        Returns:
            dict: Categories that could be aggregated
        """
        coeff_analysis = self.analyze_coefficient_similarity()
        
        aggregation_candidates = {
            'low_variation': [],  # Categories with low coefficient variation
            'unused': [],         # Categories not used in any model
            'rarely_used': [],    # Categories used in very few models
            'similar_pairs': []   # Pairs of categories with similar coefficients
        }
        
        # Find categories with low variation (could be aggregated)
        for category, analysis in coeff_analysis.items():
            if analysis['coefficient_of_variation'] < similarity_threshold:
                aggregation_candidates['low_variation'].append({
                    'category': category,
                    'cv': analysis['coefficient_of_variation'],
                    'mean': analysis['mean'],
                    'count': analysis['count']
                })
        
        # Find unused categories
        all_categories = set(self.employment_categories.keys())
        used_categories = set(self.category_usage.keys())
        aggregation_candidates['unused'] = list(all_categories - used_categories)
        
        # Find rarely used categories (used in fewer than 3 contexts)
        for category, usages in self.category_usage.items():
            if len(usages) < 3:
                aggregation_candidates['rarely_used'].append({
                    'category': category,
                    'usage_count': len(usages),
                    'contexts': usages
                })
        
        return aggregation_candidates
    
    def generate_category_analysis(self):
        """
        Generate detailed analysis of employment category usage and aggregation potential.
        """
        print("\n" + "=" * 80)
        print("EMPLOYMENT CATEGORY ANALYSIS & AGGREGATION RECOMMENDATIONS")
        print("=" * 80)
        
        # Category usage summary
        print("\n📊 EMPLOYMENT CATEGORY USAGE SUMMARY")
        print("-" * 50)
        
        used_categories = len(self.category_usage)
        total_categories = len(self.employment_categories)
        
        print(f"Categories found in models: {used_categories}/{total_categories}")
        print(f"Category utilization rate: {used_categories/total_categories:.1%}")
        
        # Most and least used categories
        category_counts = {cat: len(usages) for cat, usages in self.category_usage.items()}
        sorted_categories = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)
        
        print(f"\nMost used categories:")
        for category, count in sorted_categories[:10]:
            print(f"  • {category}: {count} occurrences")
        
        # Model-specific category usage
        print(f"\n🏗️ MODEL-SPECIFIC CATEGORY USAGE")
        print("-" * 50)
        
        model_categories = defaultdict(set)
        for category, usages in self.category_usage.items():
            for usage in usages:
                model_file = usage['file'].replace('.xls', '').replace('.xlsx', '')
                model_categories[model_file].add(category)
        
        for model, categories in sorted(model_categories.items()):
            print(f"\n{model} uses {len(categories)} employment categories:")
            for category in sorted(categories):
                usage_count = len(self.category_usage[category])
                print(f"  • {category} ({usage_count} total uses)")
        
        # Aggregation analysis
        print(f"\n🔄 AGGREGATION POTENTIAL ANALYSIS")
        print("-" * 50)
        
        candidates = self.identify_aggregation_candidates()
        
        print(f"Unused categories ({len(candidates['unused'])}):")
        for category in sorted(candidates['unused']):
            print(f"  • {category}")
        
        print(f"\nRarely used categories ({len(candidates['rarely_used'])}):")
        for item in sorted(candidates['rarely_used'], key=lambda x: x['usage_count']):
            print(f"  • {item['category']}: {item['usage_count']} uses")
        
        if candidates['low_variation']:
            print(f"\nCategories with similar coefficients (potential for aggregation):")
            for item in sorted(candidates['low_variation'], key=lambda x: x['cv']):
                print(f"  • {item['category']}: CV={item['cv']:.3f}, mean={item['mean']:.3f}, n={item['count']}")
        
        # Recommendations
        print(f"\n💡 AGGREGATION RECOMMENDATIONS")
        print("-" * 50)
        
        if candidates['unused']:
            print(f"1. REMOVE UNUSED CATEGORIES ({len(candidates['unused'])} categories)")
            print(f"   Categories never referenced in any model:")
            for category in sorted(candidates['unused']):
                print(f"   - {category}")
        
        if candidates['rarely_used']:
            print(f"\n2. CONSIDER AGGREGATING RARELY USED CATEGORIES ({len(candidates['rarely_used'])} categories)")
            print(f"   Categories used in fewer than 3 contexts:")
            for item in sorted(candidates['rarely_used'], key=lambda x: x['usage_count']):
                print(f"   - {item['category']} (only {item['usage_count']} uses)")
        
        if candidates['low_variation']:
            print(f"\n3. AGGREGATE CATEGORIES WITH SIMILAR COEFFICIENTS ({len(candidates['low_variation'])} categories)")
            print(f"   Categories with low coefficient variation (similar behavior):")
            for item in sorted(candidates['low_variation'], key=lambda x: x['cv']):
                print(f"   - {item['category']} (CV: {item['cv']:.3f})")
        
        # Calculate potential reduction
        total_removable = len(candidates['unused']) + len(candidates['rarely_used']) + len(candidates['low_variation'])
        reduction_potential = total_removable / total_categories
        
        print(f"\n📈 SUMMARY")
        print(f"   Current categories: {total_categories}")
        print(f"   Actually used: {used_categories}")
        print(f"   Potential for reduction: {total_removable} categories ({reduction_potential:.1%})")
        print(f"   Recommended target: {total_categories - total_removable} categories")
        
        # Save detailed category analysis
        self.save_category_analysis(candidates)
    
    def generate_report(self):
        """
        Generate and print a summary report of employment-related findings.
        """
        print("\n" + "=" * 60)
        print("EMPLOYMENT VARIABLE SUMMARY REPORT")
        print("=" * 60)
        
        if not self.results:
            print("No employment-related data found in any Excel files.")
            return
        
        # Overall statistics
        total_files_with_emp = len(self.results)
        total_sheets_with_emp = sum(len(sheets) for sheets in self.results.values())
        total_findings = sum(
            len(sheet['findings']) 
            for sheets in self.results.values() 
            for sheet in sheets
        )
        
        print(f"Total files with employment data: {total_files_with_emp}")
        print(f"Total sheets with employment data: {total_sheets_with_emp}")
        print(f"Total employment-related findings: {total_findings}")
        print()
        
        # Keyword frequency analysis
        keyword_counts = defaultdict(int)
        for sheets in self.results.values():
            for sheet in sheets:
                for finding in sheet['findings']:
                    for keyword in finding['keywords']:
                        keyword_counts[keyword] += 1
        
        print("Most common employment keywords found:")
        sorted_keywords = sorted(keyword_counts.items(), key=lambda x: x[1], reverse=True)
        for keyword, count in sorted_keywords[:15]:  # Top 15
            print(f"  {keyword}: {count} occurrences")
        print()
        
        # Detailed findings by file
        print("DETAILED FINDINGS BY FILE:")
        print("-" * 40)
        
        for file_name, sheets in sorted(self.results.items()):
            print(f"\n📁 {file_name}")
            for sheet_data in sheets:
                sheet_name = sheet_data['sheet_name']
                findings = sheet_data['findings']
                print(f"  📊 Sheet: {sheet_name} ({len(findings)} findings)")
                
                # Group findings by type
                by_type = defaultdict(list)
                for finding in findings:
                    by_type[finding['type']].append(finding)
                
                for finding_type, type_findings in by_type.items():
                    print(f"    {finding_type.replace('_', ' ').title()}:")
                    for finding in type_findings[:5]:  # Show first 5 of each type
                        keywords_str = ', '.join(finding['keywords'])
                        categories_str = ', '.join(finding.get('categories', []))
                        category_info = f" [Categories: {categories_str}]" if categories_str else ""
                        print(f"      • {finding['location']}: '{finding['content'][:50]}...' (keywords: {keywords_str}){category_info}")
                    if len(type_findings) > 5:
                        print(f"      ... and {len(type_findings) - 5} more")
        
        # Save detailed report to file
        self.save_detailed_report()
    
    def save_detailed_report(self):
        """
        Save a detailed report to a text file.
        """
        report_path = self.directory_path / "employment_variable_analysis_report.txt"
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("EMPLOYMENT VARIABLE ANALYSIS REPORT\n")
            f.write("=" * 50 + "\n")
            f.write(f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Directory analyzed: {self.directory_path}\n\n")
            
            # Summary statistics
            total_files_with_emp = len(self.results)
            total_sheets_with_emp = sum(len(sheets) for sheets in self.results.values())
            total_findings = sum(
                len(sheet['findings']) 
                for sheets in self.results.values() 
                for sheet in sheets
            )
            
            f.write(f"Total files with employment data: {total_files_with_emp}\n")
            f.write(f"Total sheets with employment data: {total_sheets_with_emp}\n")
            f.write(f"Total employment-related findings: {total_findings}\n\n")
            
            # Detailed findings
            for file_name, sheets in sorted(self.results.items()):
                f.write(f"\nFILE: {file_name}\n")
                f.write("-" * 30 + "\n")
                
                for sheet_data in sheets:
                    sheet_name = sheet_data['sheet_name']
                    findings = sheet_data['findings']
                    f.write(f"\nSheet: {sheet_name}\n")
                    f.write(f"Findings: {len(findings)}\n")
                    
                    for i, finding in enumerate(findings, 1):
                        f.write(f"\n  {i}. Location: {finding['location']}\n")
                        f.write(f"     Content: {finding['content']}\n")
                        f.write(f"     Keywords: {', '.join(finding['keywords'])}\n")
                        if finding.get('categories'):
                            f.write(f"     Categories: {', '.join(finding['categories'])}\n")
                        if finding.get('coefficients'):
                            f.write(f"     Coefficients: {', '.join(map(str, finding['coefficients']))}\n")
                        f.write(f"     Type: {finding['type']}\n")
        
        print(f"\nDetailed report saved to: {report_path}")
    
    def save_category_analysis(self, candidates):
        """
        Save detailed category analysis to a separate file.
        
        Args:
            candidates (dict): Aggregation candidates analysis
        """
        analysis_path = self.directory_path / "employment_category_aggregation_analysis.txt"
        
        with open(analysis_path, 'w', encoding='utf-8') as f:
            f.write("EMPLOYMENT CATEGORY AGGREGATION ANALYSIS\n")
            f.write("=" * 60 + "\n")
            f.write(f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Directory analyzed: {self.directory_path}\n\n")
            
            # Category usage by model
            f.write("CATEGORY USAGE BY MODEL\n")
            f.write("-" * 30 + "\n\n")
            
            model_categories = defaultdict(set)
            for category, usages in self.category_usage.items():
                for usage in usages:
                    model_file = usage['file'].replace('.xls', '').replace('.xlsx', '')
                    model_categories[model_file].add(category)
            
            for model, categories in sorted(model_categories.items()):
                f.write(f"{model}:\n")
                f.write(f"  Uses {len(categories)} employment categories:\n")
                for category in sorted(categories):
                    usage_count = len(self.category_usage[category])
                    f.write(f"    - {category} ({usage_count} total occurrences)\n")
                f.write("\n")
            
            # Detailed usage for each category
            f.write("\nDETAILED CATEGORY USAGE\n")
            f.write("-" * 30 + "\n\n")
            
            for category in sorted(self.category_usage.keys()):
                usages = self.category_usage[category]
                f.write(f"{category.upper()} ({len(usages)} occurrences):\n")
                
                # Group by model
                by_model = defaultdict(list)
                for usage in usages:
                    model = usage['file'].replace('.xls', '').replace('.xlsx', '')
                    by_model[model].append(usage)
                
                for model, model_usages in sorted(by_model.items()):
                    f.write(f"  {model} ({len(model_usages)} uses):\n")
                    for usage in model_usages:
                        f.write(f"    - Sheet: {usage['sheet']}, {usage['location']}\n")
                        f.write(f"      Content: {usage['content'][:80]}...\n")
                f.write("\n")
            
            # Aggregation recommendations
            f.write("AGGREGATION RECOMMENDATIONS\n")
            f.write("-" * 30 + "\n\n")
            
            f.write("1. UNUSED CATEGORIES (Remove completely):\n")
            for category in sorted(candidates['unused']):
                f.write(f"   - {category}\n")
            f.write(f"   Total: {len(candidates['unused'])} categories\n\n")
            
            f.write("2. RARELY USED CATEGORIES (Consider aggregating):\n")
            for item in sorted(candidates['rarely_used'], key=lambda x: x['usage_count']):
                f.write(f"   - {item['category']}: {item['usage_count']} uses\n")
                for usage in item['contexts']:
                    model = usage['file'].replace('.xls', '').replace('.xlsx', '')
                    f.write(f"     • {model}: {usage['sheet']}\n")
            f.write(f"   Total: {len(candidates['rarely_used'])} categories\n\n")
            
            if candidates['low_variation']:
                f.write("3. CATEGORIES WITH SIMILAR COEFFICIENTS:\n")
                for item in sorted(candidates['low_variation'], key=lambda x: x['cv']):
                    f.write(f"   - {item['category']}: CV={item['cv']:.3f}, mean={item['mean']:.3f}\n")
                f.write(f"   Total: {len(candidates['low_variation'])} categories\n\n")
            
            # Summary statistics
            total_categories = len(self.employment_categories)
            used_categories = len(self.category_usage)
            total_removable = len(candidates['unused']) + len(candidates['rarely_used']) + len(candidates['low_variation'])
            
            f.write("SUMMARY STATISTICS\n")
            f.write("-" * 20 + "\n")
            f.write(f"Total defined categories: {total_categories}\n")
            f.write(f"Categories actually used: {used_categories}\n")
            f.write(f"Utilization rate: {used_categories/total_categories:.1%}\n")
            f.write(f"Categories for potential removal/aggregation: {total_removable}\n")
            f.write(f"Recommended final category count: {total_categories - total_removable}\n")
            f.write(f"Potential reduction: {total_removable/total_categories:.1%}\n")
        
        print(f"Category analysis saved to: {analysis_path}")


def main():
    """
    Main function to run the employment variable analysis.
    """
    # Set the directory path
    directory_path = r"C:\GitHub\travel-model-two\model-files\model"
    
    # Create and run the analyzer
    analyzer = EmploymentVariableSummarizer(directory_path)
    analyzer.run_analysis()


if __name__ == "__main__":
    main()
